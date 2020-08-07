import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime  
from datetime import timedelta  

class TimeSheetWorkbook:

    def __init__(self, timesheetWorkBook ):
        self.wb = load_workbook(timesheetWorkBook)
        self.gActivityList = self.__GetNamedList('activities')
        self.gMemberList = self.__GetNamedList('members')
        self.gIterationLength = self.__GetNamedCell('iteration_length')
        self.gIteration_number = self.__GetNamedCell('iteration_number')
        self.gWork_stream = self.__GetNamedCell('work_stream')
        self.gStartDate = self.__GetNamedCell('start_date')
        self.gITERATION_HEADER_COUNT = 4
        self.gMeltedData = []

    def __GetNamedList( self, listname ):
        my_range = self.wb.defined_names[listname]
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
        cells = []
        for title, coord in dests:
            ws = self.wb[title]
            cells.append(ws[coord])
        activities = [] 
        for row in cells[0]:
            if row[0].value:
                activities.append(row[0].value)
        return(activities)


    def __GetNamedCell( self, listname ):
        retval = ""
        my_range = self.wb.defined_names[listname]
        dests = list(my_range.destinations) # returns a generator of (worksheet title, cell range) tuples
        for title, coord in dests:
            ws = self.wb[title]
            retval = ws[coord].value
        return(retval)

    def __MeltMemberData(self, sheetName ):
        ws = self.wb[sheetName]
        #Loop for the number of days
        
        for colindex in range(self.gIterationLength):
            column = self.gITERATION_HEADER_COUNT + colindex
            #day = ws.cell(1,column).value
            day = self.gStartDate + timedelta(days=colindex)
            nActivities = len(self.gActivityList)
            for rowindex in range(nActivities):
                activity = self.gActivityList[rowindex]
                val = ws.cell(2+rowindex,column).value
                if val:
                    if type(val) != str:
                        self.gMeltedData.append({'WorkStream': self.gWork_stream, 'Iteration': self.gIteration_number, 'Member':sheetName ,'Day':day,  'Activity':activity,'Effort':val})
                        #self.gMeltedData.append({'Team':team , 'WorkStream': workstream, 'Iteration': 1, 'Member': name, 'Day': self.gStartDate + timedelta(days=day), 'Activity':activity, 'Effort': effort})
            
    def GetMeltedData(self):
        self.gMeltedData = []
        nMembers = len(self.gMemberList)
        for memberId in range(nMembers):
            self.__MeltMemberData( self.gMemberList[memberId] )
        return(self.gMeltedData)


