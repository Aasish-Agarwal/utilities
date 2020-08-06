import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime  
from datetime import timedelta  

# Read named range Coordinators
# Read from row number 2
# for each row
# Push data for each date in a dictionary


class CoordinatorWorkbook:

    def __init__(self, CoordinatorsMetaWorkBook ):
        self.wb = load_workbook(CoordinatorsMetaWorkBook)
        self.gIterationLength = self.__GetNamedCell('iteration_length')
        self.gStartDate = self.__GetNamedCell('start_date')
        self.gMeltedData = []

    def __GetNamedList( self, listname ):
        my_range = self.wb.defined_names[listname]
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
        cells = []
        for title, coord in dests:
            ws = self.wb[title]
            cells.append(ws[coord])
        activities = [] 
        for row in cells[0]:
            if row[0].value:
                activities.append(row[0].value)
        return(activities)


    def __GetNamedCell( self, listname ):
        retval = ""
        my_range = self.wb.defined_names[listname]
        dests = list(my_range.destinations) # returns a generator of (worksheet title, cell range) tuples
        for title, coord in dests:
            ws = self.wb[title]
            retval = ws[coord].value
        return(retval)

    def __AppendDailyData(self , name, workstream, team, activity, allocation):
        #Team	WorkStream	Iteration	Member	Day	Activity	Effort
        effort = 8.0 * allocation / 100
        for day in range(self.gIterationLength):
            weekno = (self.gStartDate + timedelta(days=day)).weekday()
            if weekno < 5:
                self.gMeltedData.append({'Team':team , 'WorkStream': workstream, 'Iteration': 1, 'Member': name, 'Day': self.gStartDate + timedelta(days=day), 'Activity':activity, 'Effort': effort})

            
    def __MeltMemberData(self ):
        #Loop for the number of days
        my_range = self.wb.defined_names["Coordinators"]
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples

        cells = []
        for title, coord in dests:
            ws = self.wb[title]
            cells.append(ws[coord])
        activities = []
        currentRow = 1;
        for row in cells[0]:
            if currentRow < 2:
                currentRow = currentRow + 1
                continue
            if row[0].value:
                name = row[0].value
                workstream = row[1].value
                team = row[2].value
                activity = row[3].value
                allocation = row[4].value
                self.__AppendDailyData(name, workstream, team, activity, allocation)

                
    def GetMeltedData(self):
        self.gMeltedData = []
        self.__MeltMemberData()
        return(self.gMeltedData)
