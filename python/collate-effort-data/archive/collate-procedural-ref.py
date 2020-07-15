# Read activities
# Read members
# Read each member sheet
# Read data for each date
# Create a data frame - WS, iteration, team, member, date, activity , value

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('Team Efforts Distribution Recording.xlsm')

def GetNamedList( listname ):
    my_range = wb.defined_names[listname]
    dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
    cells = []
    for title, coord in dests:
        ws = wb[title]
        cells.append(ws[coord])
    activities = [] 
    for row in cells[0]:
        if row[0].value:
            activities.append(row[0].value)
    return(activities)


def GetNamedCell( listname ):
    retval = ""
    my_range = wb.defined_names[listname]
    dests = list(my_range.destinations) # returns a generator of (worksheet title, cell range) tuples
    for title, coord in dests:
        ws = wb[title]
        retval = ws[coord].value
    return(retval)

def DumpSheet( sheetName ):
    ws = wb[sheetName]
    #Loop for the number of days
    for colindex in range(gIterationLength):
        column = gITERATION_HEADER_COUNT + colindex
        day = ws.cell(1,column).value
        #print(day)
        nActivities = len(gActivityList)
        for rowindex in range(nActivities):
            val = ws.cell(2+rowindex,column).value
            #print("\t", 2+rowindex,",",column, "\t", gActivityList[rowindex])
            if val:
                print( gWork_stream, "," , sheetName , ",", day, "," , gActivityList[rowindex], "," , val)
        
def DumpWorkStream():
    nMembers = len(gMemberList)
    for memberId in range(nMembers):
        DumpSheet( gMemberList[memberId] )

  
gActivityList = GetNamedList('activities')
gMemberList = GetNamedList('members')
gIterationLength = GetNamedCell('iteration_length')
gIteration_number = GetNamedCell('iteration_number')
gWork_stream = GetNamedCell('work_stream')
gITERATION_HEADER_COUNT = 4

DumpWorkStream()




#exec(open('collate.py').read())
