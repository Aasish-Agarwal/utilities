import sys

from openpyxl import Workbook
from openpyxl import load_workbook

class TimeSheetWorkbook:

    def __init__(self, timesheetWorkBook ):
        self.wb = load_workbook(timesheetWorkBook)
        self.gActivityList = self.__GetNamedList('activities')
        self.gMemberList = self.__GetNamedList('members')
        self.gIterationLength = self.__GetNamedCell('iteration_length')
        self.gIteration_number = self.__GetNamedCell('iteration_number')
        self.gWork_stream = self.__GetNamedCell('work_stream')
        self.gITERATION_HEADER_COUNT = 4
        self.gMeltedData = []

    def __GetNamedList( self, listname ):
        my_range = self.wb.defined_names[listname]
        dests = my_range.destinations # returns a generator of (worksheet title, cell range) tuples
        cells = []
        for title, coord in dests:
            ws = self.wb[title]
            cells.append(ws[coord])
        activities = [] 
        for row in cells[0]:
            if row[0].value:
                activities.append(row[0].value)
        return(activities)


    def __GetNamedCell( self, listname ):
        retval = ""
        my_range = self.wb.defined_names[listname]
        dests = list(my_range.destinations) # returns a generator of (worksheet title, cell range) tuples
        for title, coord in dests:
            ws = self.wb[title]
            retval = ws[coord].value
        return(retval)

    def __MeltMemberData(self, sheetName ):
        ws = self.wb[sheetName]
        #Loop for the number of days
        categories = {}
        categories['Automation'] = 'Innovation'
        categories['Billable'] = 'Innovation'
        categories['Compliance'] = 'Innovation'
        categories['Cust Escalation'] = 'Innovation'
        categories['DevOps Mgmt.'] = 'Innovation'
        categories['Functional Testing'] = 'Innovation'
        categories['Inst Mgmt.'] = 'Mainteance'
        categories['L & D'] = 'Mainteance'
        categories['Leaves'] = 'Mainteance'
        categories['MIS'] = 'Mainteance'
        categories['Others'] = 'Mainteance'
        categories['Paid FE'] = 'Mainteance'
        categories['Pref Mgmt.'] = 'Mainteance'
        categories['R & D'] = 'Billable'
        categories['Recon Mgmt.'] = 'Billable'
        categories['Release Mgmt.'] = 'Billable'
        categories['Roadmap FE'] = 'Billable'
        categories['Scrum/Kanban Meetings'] = 'Billable'
        categories['Stop & Fix'] = 'Keeping Lights On'
        categories['Tools'] = 'Keeping Lights On'
        categories['Upgrade Support'] = 'Keeping Lights On'
        
        for colindex in range(self.gIterationLength):
            column = self.gITERATION_HEADER_COUNT + colindex
            day = ws.cell(1,column).value
            nActivities = len(self.gActivityList)
            for rowindex in range(nActivities):
                activity = self.gActivityList[rowindex]
                category = categories[activity]
                val = ws.cell(2+rowindex,column).value
                if val:
                    self.gMeltedData.append({'WorkStream': self.gWork_stream, 'Iteration': self.gIteration_number, 'Member':sheetName ,'Day':day, 'Category': category, 'Activity':activity,'Effort':val})
            
    def GetMeltedData(self):
        self.gMeltedData = []
        nMembers = len(self.gMemberList)
        for memberId in range(nMembers):
            self.__MeltMemberData( self.gMemberList[memberId] )
        return(self.gMeltedData)


#exec(open('TimeSheetWokbook.py').read())

